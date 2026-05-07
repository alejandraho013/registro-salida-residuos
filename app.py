import streamlit as st
import pandas as pd
from github import Github
from datetime import datetime
import re
import io
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────
REPO_NAME = "alejandraho013/registro-salida-residuos"
    
GESTORES_DATA = {
    "CORPOGESTAR":              sorted(["Cartón limpio","Cartón contaminado","Papel de archivo","Pasta","PET limpio","PET contaminado","Plástico","Retal de tela","Tubo plega"]),
    "Recicla Oriente":          sorted(["Cartón limpio","Cartón contaminado","Papel de archivo","Pasta","PET limpio","PET contaminado","Plástico","Retal de tela","Tubo plega"]),
    "VEOLIA Aprovechables": sorted(["Algodón","Retal de tela","Tubo plega","CDR"]),
    "VEOLIA Peligrosos":    sorted(["RAEE","Residuos laboratorio","CDR","Solidos Contaminados"]),
    "Otro": [],
}

COLORES_EMPRESA = {
    "CORPOGESTAR":              "#2196F3",
    "Recicla Oriente":          "#4CAF50",
    "VEOLIA Aprovechables": "#FF9800",
    "VEOLIA Peligrosos":    "#F44336",
}

# Colores openpyxl para cada gestor (ARGB sin #)
COLORES_GESTOR_XL = {
    "CORPOGESTAR":              "FF2196F3",
    "Recicla Oriente":          "FF4CAF50",
    "VEOLIA Aprovechables": "FFFF9800",
    "VEOLIA Peligrosos":    "FFF44336",
}
COLOR_HEADER_XL   = "FF1A237E"   # azul oscuro encabezados
COLOR_SUBHEAD_XL  = "FFE8EAF6"   # lila muy claro filas alternas / sub-encabezados

try:
    TOKEN     = st.secrets["TOKEN"]
    REPO_NAME = st.secrets.get("REPO_NAME", REPO_NAME)
except Exception:
    st.error("⚠️ Configura el TOKEN en los Secrets de Streamlit Cloud.")
    st.stop()

# ─────────────────────────────────────────────────────────────
# ESTILOS Y SESSION STATE
# ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="TINTATEX · Gestión de Residuos", layout="wide")

st.markdown("""
<style>
.main-header {
    background: linear-gradient(135deg,#1a237e 0%,#283593 100%);
    padding:1.5rem; border-radius:12px; color:white; margin-bottom:1.5rem;
}
.kpi-card {
    background:white; border-radius:10px; padding:1rem;
    border-left:4px solid #1a237e;
    box-shadow:0 2px 5px rgba(0,0,0,.1); text-align:center;
}
.kpi-value { font-size:1.8rem; font-weight:700; color:#1a237e; }
.kpi-label { font-size:0.9rem; color:#666; }
</style>
<div class="main-header">
  <h1>🏭 TINTATEX · Gestión de Residuos</h1>
  <p>Registro de Pesajes y Control de Carga</p>
</div>
""", unsafe_allow_html=True)

for key, default in [("lista_temporal", []), ("envio_exitoso", False)]:
    if key not in st.session_state:
        st.session_state[key] = default

# ─────────────────────────────────────────────────────────────
# COLUMNAS ESTÁNDAR
# ─────────────────────────────────────────────────────────────
COLUMNAS_MASTER = ["fecha","mes","empresa","conductor","placa","tipo_residuo","peso_kg","novedades"]

# ─────────────────────────────────────────────────────────────
# FUNCIONES DE DATOS — GITHUB
# ─────────────────────────────────────────────────────────────
def _get_repo():
    return Github(TOKEN).get_repo(REPO_NAME)


@st.cache_data(ttl=120)
def cargar_datos_github() -> pd.DataFrame:
    try:
        repo = _get_repo()
        try:
            f = repo.get_contents("database.xlsx")
        except Exception:
            return pd.DataFrame(columns=COLUMNAS_MASTER)

        df = pd.read_excel(
            io.BytesIO(f.decoded_content),
            sheet_name="MASTER",
            engine="openpyxl",
        )
        df["fecha"] = pd.to_datetime(df.get("fecha"), errors="coerce")
        if "mes" not in df.columns or df["mes"].isna().all():
            df["mes"] = df["fecha"].dt.strftime("%B_%Y")
        if "peso_kg" in df.columns:
            df["peso_kg"] = pd.to_numeric(df["peso_kg"], errors="coerce").fillna(0)
        return df

    except Exception as e:
        st.error(f"❌ Error al cargar datos desde GitHub: {e}")
        return pd.DataFrame(columns=COLUMNAS_MASTER)


def _nombre_hoja(empresa: str) -> str:
    nombre = re.sub(r'[\\/*?:\[\]]', "", empresa).strip()[:31]
    return nombre.upper() if nombre else "OTROS"


def guardar_datos(nuevas_filas: list, empresa: str, placa: str):
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    repo = _get_repo()

    try:
        xlsx_file = repo.get_contents("database.xlsx")
        dicc = pd.read_excel(
            io.BytesIO(xlsx_file.decoded_content),
            sheet_name=None,
            engine="openpyxl",
        )
        sha = xlsx_file.sha
    except Exception:
        dicc = {}
        sha  = None

    df_nuevos    = pd.DataFrame(nuevas_filas)
    nombre_hoja  = _nombre_hoja(empresa)

    dicc["MASTER"] = pd.concat(
        [dicc.get("MASTER", pd.DataFrame(columns=COLUMNAS_MASTER)), df_nuevos],
        ignore_index=True,
    )
    dicc[nombre_hoja] = pd.concat(
        [dicc.get(nombre_hoja, pd.DataFrame(columns=COLUMNAS_MASTER)), df_nuevos],
        ignore_index=True,
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for hoja, datos in dicc.items():
            datos.to_excel(writer, sheet_name=hoja, index=False)

    commit_msg = f"Reg_{empresa}_{placa}_{ts}"
    if sha:
        repo.update_file("database.xlsx", commit_msg, output.getvalue(), sha)
    else:
        repo.create_file("database.xlsx", commit_msg, output.getvalue())

    cargar_datos_github.clear()


# ─────────────────────────────────────────────────────────────
# GENERADOR DE EXCEL DE DESCARGA (Resumen + MASTER + gestores)
# ─────────────────────────────────────────────────────────────
def _estilo_header(ws, fila: int, n_cols: int, color_hex: str = COLOR_HEADER_XL):
    fill   = PatternFill("solid", fgColor=color_hex)
    fuente = Font(bold=True, color="FFFFFFFF", name="Arial", size=11)
    borde  = Border(
        bottom=Side(style="medium", color="FFFFFFFF"),
    )
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill   = fill
        cell.font   = fuente
        cell.border = borde
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws):
    anchos = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            col_idx = getattr(cell, "column", None)
            if not isinstance(col_idx, int) or cell.value is None:
                continue
            largo = len(str(cell.value))
            if largo > anchos.get(col_idx, 0):
                anchos[col_idx] = largo
    for col_idx, largo in anchos.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(largo + 4, 40)


@st.cache_data(ttl=300, show_spinner=False)
def construir_excel_descarga(df_filtrado: pd.DataFrame) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    FUENTE_BASE = Font(name="Arial", size=10)
    FILL_ALT    = PatternFill("solid", fgColor="FFF5F5F5")

    def _escribir_tabla(ws, df_data, encabezados, fila_ini=1, col_ini=1, color_hdr=COLOR_HEADER_XL):
        """Escribe encabezados + filas con filas alternas y devuelve fila siguiente."""
        for j, h in enumerate(encabezados, start=col_ini):
            ws.cell(row=fila_ini, column=j, value=h)
        _estilo_header(ws, fila_ini, len(encabezados) + col_ini - 1, color_hdr)

        for i, row in enumerate(df_data.itertuples(index=False), start=fila_ini + 1):
            fill_row = FILL_ALT if i % 2 == 0 else PatternFill()
            for j, val in enumerate(row, start=col_ini):
                cell = ws.cell(row=i, column=j, value=val)
                cell.font = FUENTE_BASE
                cell.fill = fill_row
                cell.alignment = Alignment(horizontal="center")
        return fila_ini + len(df_data) + 1

    cols_mostrar  = ["fecha","mes","empresa","conductor","placa","tipo_residuo","peso_kg","novedades"]
    encabezados_m = ["Fecha","Mes","Empresa","Conductor","Placa","Tipo Residuo","Peso (kg)","Novedades"]

    # ── 1. HOJA RESUMEN ─────────────────────────────────────
    ws_s = wb.create_sheet("Resumen")

    ws_s.merge_cells("A1:H1")
    title_s = ws_s["A1"]
    title_s.value     = "RESUMEN EJECUTIVO · TINTATEX"
    title_s.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=14)
    title_s.fill      = PatternFill("solid", fgColor=COLOR_HEADER_XL)
    title_s.alignment = Alignment(horizontal="center", vertical="center")
    ws_s.row_dimensions[1].height = 28

    kpis = [
        ("Peso total (kg)",        f"{df_filtrado['peso_kg'].sum():,.1f}"),
        ("Registros",              len(df_filtrado)),
        ("Empresas",               df_filtrado["empresa"].nunique()),
        ("Tipos de residuo",       df_filtrado["tipo_residuo"].nunique()),
        ("Promedio / registro",    f"{df_filtrado['peso_kg'].mean():,.1f} kg"),
    ]
    for i, (lbl, val) in enumerate(kpis, start=1):
        c_lbl = ws_s.cell(row=3, column=i, value=lbl)
        c_lbl.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
        c_lbl.fill      = PatternFill("solid", fgColor=COLOR_HEADER_XL)
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")
        c_val = ws_s.cell(row=4, column=i, value=val)
        c_val.font      = Font(bold=True, name="Arial", size=12, color="FF1A237E")
        c_val.fill      = PatternFill("solid", fgColor=COLOR_SUBHEAD_XL)
        c_val.alignment = Alignment(horizontal="center", vertical="center")
    ws_s.row_dimensions[4].height = 22

    # Tabla cruzada empresa × residuo con totales en fila y columna
    pivot = (
        df_filtrado.pivot_table(
            index="empresa", columns="tipo_residuo",
            values="peso_kg", aggfunc="sum", fill_value=0,
        ).round(2)
    )
    pivot["TOTAL"]     = pivot.sum(axis=1).round(2)
    pivot.loc["TOTAL"] = pivot.sum(axis=0).round(2)
    pivot.columns.name = None
    pivot_reset        = pivot.reset_index().rename(columns={"empresa": "Empresa"})

    fila_cruce = 7
    ws_s.cell(row=fila_cruce - 1, column=1, value="Cruce Empresa × Residuo (kg)").font = Font(bold=True, name="Arial", size=11, color="FF1A237E")
    _escribir_tabla(ws_s, pivot_reset, list(pivot_reset.columns), fila_ini=fila_cruce)

    n_filas_pivot = len(pivot_reset)
    n_cols_pivot  = len(pivot_reset.columns)
    fila_total_p  = fila_cruce + n_filas_pivot

    # Destacar fila TOTAL (última) y columna TOTAL (última)
    for j in range(1, n_cols_pivot + 1):
        c = ws_s.cell(row=fila_total_p, column=j)
        c.font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER_XL)
        c.alignment = Alignment(horizontal="center")
    for i in range(fila_cruce + 1, fila_total_p):
        c = ws_s.cell(row=i, column=n_cols_pivot)
        c.font = Font(bold=True, name="Arial", size=10, color="FF1A237E")
        c.fill = PatternFill("solid", fgColor=COLOR_SUBHEAD_XL)

    # Tablas auxiliares para gráficas (debajo del cruce, en columnas A-B y D-E)
    resumen_emp = (
        df_filtrado.groupby("empresa")["peso_kg"].sum()
        .reset_index().sort_values("peso_kg", ascending=False)
    )
    resumen_res = (
        df_filtrado.groupby("tipo_residuo")["peso_kg"].sum()
        .reset_index().sort_values("peso_kg", ascending=False)
    )
    fila_aux = fila_total_p + 3

    ws_s.cell(row=fila_aux, column=1, value="Empresa").font   = Font(bold=True, name="Arial")
    ws_s.cell(row=fila_aux, column=2, value="Peso (kg)").font = Font(bold=True, name="Arial")
    for i, row in enumerate(resumen_emp.itertuples(index=False), start=fila_aux + 1):
        ws_s.cell(row=i, column=1, value=row.empresa)
        ws_s.cell(row=i, column=2, value=round(row.peso_kg, 2))
    n_emp = len(resumen_emp)

    ws_s.cell(row=fila_aux, column=4, value="Residuo").font   = Font(bold=True, name="Arial")
    ws_s.cell(row=fila_aux, column=5, value="Peso (kg)").font = Font(bold=True, name="Arial")
    for i, row in enumerate(resumen_res.itertuples(index=False), start=fila_aux + 1):
        ws_s.cell(row=i, column=4, value=row.tipo_residuo)
        ws_s.cell(row=i, column=5, value=round(row.peso_kg, 2))
    n_res = len(resumen_res)

    # Gráfica barras (empresa)
    bar_chart = BarChart()
    bar_chart.type         = "col"
    bar_chart.title        = "Peso total por empresa (kg)"
    bar_chart.y_axis.title = "kg"
    bar_chart.x_axis.title = "Empresa"
    bar_chart.width        = 16
    bar_chart.height       = 10
    bar_chart.style        = 10
    bar_chart.add_data(
        Reference(ws_s, min_col=2, min_row=fila_aux, max_row=fila_aux + n_emp),
        titles_from_data=True,
    )
    bar_chart.set_categories(
        Reference(ws_s, min_col=1, min_row=fila_aux + 1, max_row=fila_aux + n_emp),
    )
    bar_chart.series[0].graphicalProperties.solidFill = "1A237E"
    ws_s.add_chart(bar_chart, f"G{fila_aux}")

    # Gráfica pie (residuo)
    pie_chart = PieChart()
    pie_chart.title  = "Distribución por tipo de residuo"
    pie_chart.width  = 16
    pie_chart.height = 10
    pie_chart.style  = 10
    pie_chart.add_data(
        Reference(ws_s, min_col=5, min_row=fila_aux, max_row=fila_aux + n_res),
        titles_from_data=True,
    )
    pie_chart.set_categories(
        Reference(ws_s, min_col=4, min_row=fila_aux + 1, max_row=fila_aux + n_res),
    )
    pie_chart.dataLabels = DataLabelList(showPercent=True, showCatName=True)
    ws_s.add_chart(pie_chart, f"G{fila_aux + 21}")

    _autowidth(ws_s)

    # ── 2. HOJA MASTER (ordenada por empresa, con subtotales) ──
    ws_m = wb.create_sheet("MASTER")
    ws_m.freeze_panes = "A2"

    for j, h in enumerate(encabezados_m, start=1):
        ws_m.cell(row=1, column=j, value=h)
    _estilo_header(ws_m, 1, len(encabezados_m))

    df_ord = df_filtrado.sort_values(["empresa", "fecha"])[cols_mostrar].copy()
    df_ord["fecha"] = df_ord["fecha"].astype(str)

    fila_actual   = 2
    total_general = 0.0
    n_cols_m      = len(encabezados_m)

    for empresa, grupo in df_ord.groupby("empresa", sort=True):
        color_emp = COLORES_GESTOR_XL.get(empresa, COLOR_HEADER_XL)

        for _, row in grupo.iterrows():
            for j, col in enumerate(cols_mostrar, start=1):
                cell = ws_m.cell(row=fila_actual, column=j, value=row[col])
                cell.font      = FUENTE_BASE
                cell.fill      = FILL_ALT if fila_actual % 2 == 0 else PatternFill()
                cell.alignment = Alignment(horizontal="center")
            ws_m.cell(row=fila_actual, column=7).number_format = "#,##0.0"
            fila_actual += 1

        subtotal = float(grupo["peso_kg"].sum())
        total_general += subtotal

        for j in range(1, n_cols_m + 1):
            c = ws_m.cell(row=fila_actual, column=j)
            c.fill      = PatternFill("solid", fgColor=color_emp)
            c.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_m.cell(row=fila_actual, column=1, value=f"Subtotal — {empresa}")
        ws_m.cell(row=fila_actual, column=7, value=round(subtotal, 2)).number_format = "#,##0.0"
        fila_actual += 1

    for j in range(1, n_cols_m + 1):
        c = ws_m.cell(row=fila_actual, column=j)
        c.fill      = PatternFill("solid", fgColor=COLOR_HEADER_XL)
        c.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_m.cell(row=fila_actual, column=1, value="TOTAL GENERAL")
    ws_m.cell(row=fila_actual, column=7, value=round(total_general, 2)).number_format = "#,##0.0"
    ws_m.row_dimensions[fila_actual].height = 22

    _autowidth(ws_m)

    # ── 3. HOJA POR CADA GESTOR ─────────────────────────────
    gestores_presentes = df_filtrado["empresa"].dropna().unique()
    for gestor in sorted(gestores_presentes):
        nombre_hoja = _nombre_hoja(gestor)[:31]
        df_g = df_filtrado[df_filtrado["empresa"] == gestor][cols_mostrar].copy()
        df_g["fecha"] = df_g["fecha"].astype(str)

        color_hdr = COLORES_GESTOR_XL.get(gestor, COLOR_HEADER_XL)
        ws_g = wb.create_sheet(nombre_hoja)
        ws_g.freeze_panes = "A3"

        ws_g.merge_cells("A1:H1")
        title_cell = ws_g["A1"]
        title_cell.value     = f"REGISTROS — {gestor.upper()}"
        title_cell.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=12)
        title_cell.fill      = PatternFill("solid", fgColor=color_hdr)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_g.row_dimensions[1].height = 22

        _escribir_tabla(ws_g, df_g, encabezados_m, fila_ini=2, color_hdr=color_hdr)

        fila_tot_g = len(df_g) + 3
        for j in range(1, n_cols_m + 1):
            c = ws_g.cell(row=fila_tot_g, column=j)
            c.fill      = PatternFill("solid", fgColor=color_hdr)
            c.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_g.cell(row=fila_tot_g, column=1, value=f"TOTAL {gestor.upper()}")
        ws_g.cell(row=fila_tot_g, column=7, value=f"=SUM(G3:G{fila_tot_g-1})").number_format = "#,##0.0"

        _autowidth(ws_g)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────
# INTERFAZ
# ─────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📝 Registro", "📊 Reportes y Gráficas"])

# ── TAB 1: REGISTRO ──────────────────────────────────────────
with tab1:
    if st.session_state.envio_exitoso:
        st.success("✅ ¡Registro guardado con éxito!")
        if st.button("🔄 Nuevo Registro", type="primary"):
            st.session_state.envio_exitoso = False
            st.session_state.lista_temporal = []
            st.rerun()
    else:
        with st.expander("🚛 1. Datos del Vehículo", expanded=True):
            c1, c2 = st.columns(2)
            fecha     = c1.date_input("Fecha", datetime.now())
            emp_sel   = c1.selectbox("Empresa Gestora", options=list(GESTORES_DATA.keys()))
            empresa_final = (
                c1.text_input("Nombre Manual").upper().strip()
                if emp_sel == "Otro"
                else emp_sel
            )

            conductor = c2.text_input("Conductor")
            cp1, cp2  = c2.columns([2, 1])
            placa     = cp1.text_input("Placa (ABC123)").upper().strip()
            placa_valida = bool(re.match(r"^[A-Z]{3}[0-9]{3}$", placa))
            if placa:
                (cp2.success if placa_valida else cp2.error)(
                    "✔ Válida" if placa_valida else "✗ Inválida"
                )

            capacidad = c2.number_input("Capacidad Camión (kg) — Opcional", min_value=0.0, step=100.0)

        st.subheader("⚖️ 2. Pesajes")
        col_r, col_p, col_b = st.columns([3, 2, 1])
        res_opts  = GESTORES_DATA.get(emp_sel, []) + ["Otro"]
        tipo_res  = col_r.selectbox("Tipo de Residuo", options=res_opts)
        res_final = col_r.text_input("¿Cuál residuo?") if tipo_res == "Otro" else tipo_res
        peso      = col_p.number_input("Peso (kg)", min_value=0.0, step=0.1)

        if col_b.button("➕ Añadir", use_container_width=True):
            if not placa_valida:
                st.error("Placa inválida — formato esperado: ABC123")
            elif peso <= 0:
                st.error("El peso debe ser mayor a 0")
            else:
                st.session_state.lista_temporal.append({"tipo_residuo": res_final, "peso_kg": peso})
                st.toast(f"✅ {res_final} — {peso} kg añadido")

        if st.session_state.lista_temporal:
            df_temp  = pd.DataFrame(st.session_state.lista_temporal)
            total_kg = df_temp["peso_kg"].sum()

            if capacidad > 0 and total_kg > capacidad:
                st.warning(f"⚠️ Carga ({total_kg:,.1f} kg) supera la capacidad ({capacidad:,.1f} kg)")

            m1, m2 = st.columns(2)
            m1.metric("Suma Carga", f"{total_kg:,.1f} kg")
            m2.metric("Registros",  len(df_temp))
            st.dataframe(df_temp, use_container_width=True, hide_index=True)

            b1, b2 = st.columns(2)
            if b1.button("⏪ Eliminar último"):
                st.session_state.lista_temporal.pop()
                st.rerun()
            if b2.button("🧹 Limpiar lista"):
                st.session_state.lista_temporal = []
                st.rerun()

        novedades = st.text_area("Observaciones")

        if st.button("📤 ENVIAR TODO", type="primary", use_container_width=True):
            if not st.session_state.lista_temporal:
                st.error("Agrega al menos un pesaje antes de enviar.")
            elif not placa_valida:
                st.error("La placa no es válida.")
            elif not empresa_final:
                st.error("Escribe el nombre de la empresa gestora.")
            else:
                with st.spinner("Guardando en GitHub…"):
                    filas = [
                        {
                            "fecha":        str(fecha),
                            "mes":          fecha.strftime("%B_%Y"),
                            "empresa":      empresa_final,
                            "conductor":    conductor,
                            "placa":        placa,
                            "tipo_residuo": x["tipo_residuo"],
                            "peso_kg":      x["peso_kg"],
                            "novedades":    novedades or "Sin novedades",
                        }
                        for x in st.session_state.lista_temporal
                    ]
                    try:
                        guardar_datos(filas, empresa_final, placa)
                        st.session_state.envio_exitoso = True
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error al guardar: {e}")

# ── TAB 2: REPORTES ──────────────────────────────────────────
with tab2:
    col_ref, _ = st.columns([1, 5])
    if col_ref.button("🔄 Actualizar datos", key="refresh"):
        cargar_datos_github.clear()
        st.rerun()

    with st.spinner("Cargando datos…"):
        df = cargar_datos_github()

    if df.empty:
        st.info("📭 Sin datos registrados aún. Realiza tu primer registro en la pestaña anterior.")
    else:
        # ── FILTROS ──────────────────────────────────────────
        with st.expander("🔍 Filtros", expanded=True):
            tipo_rango = st.radio(
                "Período",
                ["Rango de fechas"],
                horizontal=True,
            )

            f1, f2, f3, f4 = st.columns([2, 2, 2, 2])

            fecha_min = df["fecha"].min().date()
            fecha_max = df["fecha"].max().date()

            if tipo_rango == "Rango de fechas":
                fecha_ini = f1.date_input("Desde", value=fecha_min, min_value=fecha_min, max_value=fecha_max)
                fecha_fin = f2.date_input("Hasta", value=fecha_max, min_value=fecha_min, max_value=fecha_max)
                mask_fecha = (df["fecha"].dt.date >= fecha_ini) & (df["fecha"].dt.date <= fecha_fin)
            else:
                mask_fecha = pd.Series([True] * len(df), index=df.index)

            empresas = ["Todas"] + sorted(df["empresa"].dropna().unique().tolist())
            emp_fil  = f3.selectbox("Empresa", empresas)
            mask_emp = (df["empresa"] == emp_fil) if emp_fil != "Todas" else pd.Series([True] * len(df), index=df.index)

            residuos = ["Todos"] + sorted(df["tipo_residuo"].dropna().unique().tolist())
            res_fil  = f4.selectbox("Tipo de residuo", residuos)
            mask_res = (df["tipo_residuo"] == res_fil) if res_fil != "Todos" else pd.Series([True] * len(df), index=df.index)

        df_f = df[mask_fecha & mask_emp & mask_res].copy()

        if df_f.empty:
            st.warning("⚠️ No hay registros con los filtros seleccionados.")
        else:
            # ── KPIs ─────────────────────────────────────────
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Peso total",          f"{df_f['peso_kg'].sum():,.1f} kg")
            k2.metric("Registros",            len(df_f))
            k3.metric("Empresas",             df_f["empresa"].nunique())
            k4.metric("Promedio / registro",  f"{df_f['peso_kg'].mean():,.1f} kg")

            st.divider()

            # ── GRÁFICAS ──────────────────────────────────────
            g1, g2 = st.columns(2)

            with g1:
                emp_agg = df_f.groupby("empresa")["peso_kg"].sum().reset_index()
                fig_bar = px.bar(
                    emp_agg, x="empresa", y="peso_kg",
                    title="Peso total por empresa",
                    color="empresa",
                    color_discrete_map=COLORES_EMPRESA,
                    labels={"peso_kg": "Peso (kg)", "empresa": ""},
                    text_auto=".0f",
                )
                fig_bar.update_traces(textposition="outside")
                fig_bar.update_layout(showlegend=False, margin=dict(t=40, b=0))
                st.plotly_chart(fig_bar, use_container_width=True)

            with g2:
                fig_pie = px.pie(
                    df_f, values="peso_kg", names="tipo_residuo",
                    title="Distribución por tipo de residuo",
                )
                fig_pie.update_traces(textposition="inside", textinfo="percent+label")
                fig_pie.update_layout(margin=dict(t=40, b=0), showlegend=False)
                st.plotly_chart(fig_pie, use_container_width=True)

            if df_f["fecha"].dt.date.nunique() > 1:
                df_tiempo = (
                    df_f.groupby(df_f["fecha"].dt.date)["peso_kg"]
                    .sum().reset_index()
                    .rename(columns={"fecha": "Fecha", "peso_kg": "Peso (kg)"})
                )
                fig_line = px.line(
                    df_tiempo, x="Fecha", y="Peso (kg)",
                    title="Evolución diaria de peso",
                    markers=True,
                )
                fig_line.update_layout(margin=dict(t=40, b=0))
                st.plotly_chart(fig_line, use_container_width=True)

            # ── TABLA DETALLE ─────────────────────────────────
            st.subheader("Detalle de registros")
            st.dataframe(
                df_f.sort_values("fecha", ascending=False).reset_index(drop=True),
                use_container_width=True,
                hide_index=True,
            )

            # ── DESCARGA ──────────────────────────────────────
            sufijo = (
                f"{emp_fil.replace(' ','_')}__"
                f"{res_fil.replace(' ','_')}__"
                f"{tipo_rango.replace(' ','_')}"
            ).replace("Todas", "todas").replace("Todos", "todos")
            nombre_archivo = f"Reporte_TINTATEX_{sufijo}.xlsx"

            with st.spinner("Preparando Excel…"):
                excel_bytes = construir_excel_descarga(df_f)

            st.download_button(
                label="⬇️ Descargar Excel",
                data=excel_bytes,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
            
